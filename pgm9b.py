from openpyxl import Workbook
from openpyxl.styles import Font

# Create a new Excel workbook and set the sheet titles
wb = Workbook()
sheet = wb.active
sheet.title = "Language"
wb.create_sheet(title="Capital")

# Define data lists
lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ['KA', 'TS', 'TN']

# Set headers and apply bold font
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Language"
sheet.cell(row=1, column=3).value = "Code"
ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft

# Populate the Language sheet with data
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i-2]
    sheet.cell(row=i, column=2).value = lang[i-2]
    sheet.cell(row=i, column=3).value = code[i-2]

# Save the workbook
wb.save("demo.xlsx")

# Access the "Capital" sheet
sheet = wb["Capital"]

# Set headers and apply bold font for the "Capital" sheet
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Capital"
sheet.cell(row=1, column=3).value = "Code"
ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft

# Populate the "Capital" sheet with data
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i-2]
    sheet.cell(row=i, column=2).value = capital[i-2]
    sheet.cell(row=i, column=3).value = code[i-2]

# Save the updated workbook
wb.save("demo.xlsx")

# User input to search for capital based on state code
srchCode = input("Enter state code for finding capital ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding capital for code", srchCode, "is", sheet.cell(row=i, column=2).value)

# Access the "Language" sheet
sheet = wb["Language"]

# User input to search for language based on state code
srchCode = input("Enter state code for finding language ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding language for code", srchCode, "is", sheet.cell(row=i, column=2).value)

# Close the workbook
wb.close()
